import openpyxl

from file_translator.services.translation_service import translate_text


async def translate_excel(
    file_path: str, source_lang="FRENCH", target_lang="English"
) -> str:
    workbook = openpyxl.load_workbook(file_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    translated_text = await translate_text(
                        cell.value, source_lang, target_lang
                    )
                    cell.value = translated_text

    workbook.save(file_path)
    return file_path
