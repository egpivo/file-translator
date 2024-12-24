import io
import os
from typing import Union

import openpyxl
from fastapi import APIRouter, HTTPException, UploadFile
from fastapi.responses import FileResponse, StreamingResponse

from file_translator.config import EXCEL_MEDIA_TYPE, SIZE_THRESHOLD
from file_translator.services.translation_service import translate_text
from file_translator.utils import make_directory, remove_directory

router = APIRouter()


async def translate_excel_in_memory(
    file: UploadFile, source_lang: str = "FRENCH", target_lang: str = "English"
) -> io.BytesIO:
    contents = await file.read()
    workbook = openpyxl.load_workbook(io.BytesIO(contents))

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    translated_text = await translate_text(
                        cell.value, source_lang, target_lang
                    )
                    cell.value = translated_text

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


async def translate_excel_on_disk(
    file: UploadFile, source_lang: str = "FRENCH", target_lang: str = "English"
) -> str:
    temp_dir = make_directory(file.filename)
    file_path = os.path.join(temp_dir, file.filename)

    try:
        with open(file_path, "wb") as f:
            f.write(await file.read())

        workbook = openpyxl.load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        translated_text = await translate_text(
                            cell.value, source_lang, target_lang
                        )
                        cell.value = translated_text

        translated_file_path = os.path.join(temp_dir, f"translated_{file.filename}")
        workbook.save(translated_file_path)
        return translated_file_path

    finally:
        remove_directory(temp_dir)


@router.post("/translate_file", response_model=None)
async def translate_excel_file(
    file: UploadFile,
) -> Union[StreamingResponse, FileResponse]:
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=400, detail="Unsupported file type. Please upload an XLSX file."
        )

    try:
        contents = await file.read()
        file_size = len(contents)
        file.file.seek(0)

        if file_size <= SIZE_THRESHOLD:
            # Process the file in memory
            translated_file = await translate_excel_in_memory(file)
            return StreamingResponse(
                translated_file,
                media_type=EXCEL_MEDIA_TYPE,
                headers={
                    "Content-Disposition": f"attachment; filename=translated_{file.filename}"
                },
            )
        else:
            # Process the file using disk storage
            translated_file_path = await translate_excel_on_disk(file)
            return FileResponse(
                path=translated_file_path,
                media_type=EXCEL_MEDIA_TYPE,
                filename=f"translated_{file.filename}",
            )
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"An error occurred during processing: {str(e)}"
        )
